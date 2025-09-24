
import streamlit as st
from PIL import Image
import io
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, Border, Side
from streamlit_webrtc import webrtc_streamer, WebRtcMode
import av


# --- Page Configuration ---
st.set_page_config(
    page_title="VZW Site Documentation",
    page_icon="✅",
    layout="centered",
)

# --- App State Management ---
if 'stage' not in st.session_state:
    st.session_state.stage = 0
if 'baseband_data' not in st.session_state:
    st.session_state.baseband_data = {}
if 'captured_images' not in st.session_state:
    st.session_state.captured_images = {}

# --- Photo List ---
# As per 'Photo List' sheet
photo_list = [
    "Pre-Installation Site Photo",
    "Sector A Pre-Installation",
    "Sector B Pre-Installation",
    "Sector C Pre-Installation",
    "Pre-Installation Cabinet/Rack Photo",
    "Pre-Installation Radio/RRH Photo",
    "Pre-Installation Antenna/Filter Photo",
    "Pre-Installation Fiber/Power Cable Photo",
    "Post-Installation Site Photo",
    "Sector A Post-Installation",
    "Sector B Post-Installation",
    "Sector C Post-Installation",
    "Post-Installation Cabinet/Rack Photo",
    "Post-Installation Radio/RRH Photo",
    "Post-Installation Antenna/Filter Photo",
    "Post-Installation Fiber/Power Cable Photo",
    "Equipment Serial Numbers",
    "Additional Photo 1",
    "Additional Photo 2",
    "Additional Photo 3",
    "Additional Photo 4",
    "Additional Photo 5",
    "Additional Photo 6",
    "Additional Photo 7",
    "Additional Photo 8",
    "Additional Photo 9",
    "Additional Photo 10",
    "Additional Photo 11",
    "Additional Photo 12",
    "Additional Photo 13",
    "Additional Photo 14",
    "Additional Photo 15",
    "Additional Photo 16",
    "Additional Photo 17",
    "Additional Photo 18",
    "Additional Photo 19",
    "Additional Photo 20",
    "Additional Photo 21",
    "Additional Photo 22",
    "Additional Photo 23",
]


def set_stage(stage):
    st.session_state.stage = stage

# --- Stage 1: Gen 4 BB Conversion Form ---
def baseband_swap_form():
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("Gen 4 BB Conversion")
    with col2:
        st.image("https://upload.wikimedia.org/wikipedia/commons/a/a7/Ericsson_logo.svg", width=150)

    with st.form("baseband_form"):
        # Two-column layout
        col1, col2 = st.columns(2)

        with col1:
            st.session_state.baseband_data['antenna_location'] = st.radio(
                "Antenna Location:",
                ('Rooftop', 'Monopole', 'Lattice Tower', 'Other'))
            if st.session_state.baseband_data['antenna_location'] == 'Other':
                st.session_state.baseband_data['antenna_location_other'] = st.text_input("Please specify other location:")

            st.session_state.baseband_data['installation'] = st.text_input("Installation:", "Gen2/3 BB removal, Gen 4 BB install")

        with col2:
            st.session_state.baseband_data['site_name'] = st.text_input("Site Name:", "PHO_GUNPOWDER")
            st.session_state.baseband_data['contractor'] = st.text_input("Contractor:", "Integer")
            st.session_state.baseband_data['tech_name'] = st.text_input("Tech Name:", "Wilfred")
            st.session_state.baseband_data['date'] = st.date_input("Date:")
            st.session_state.baseband_data['project'] = st.text_input("Project:", "Gen 4 - BB6672 Install")

        st.subheader("Additional Notes")
        st.session_state.baseband_data['additional_notes'] = st.text_area("", height=150)

        submitted = st.form_submit_button("Next: Capture Photos")
        if submitted:
            # Simple validation
            if st.session_state.baseband_data.get('site_name') and st.session_state.baseband_data.get('tech_name'):
                set_stage(1)
                st.rerun()
            else:
                st.error("Please fill in all required fields (Site Name, Tech Name).")


# --- Stage 2: Photo Capture ---
def photo_capture_screen():
    st.header("Capture Photos")

    # Dropdown for photo selection
    photo_name = st.selectbox("Select a photo to capture:", photo_list)

    camera_facing_mode = st.radio("Select Camera", ("Front", "Back"), horizontal=True, index=1) # Default to Back camera
    facing_mode = "user" if camera_facing_mode == "Front" else "environment"

    webrtc_ctx = webrtc_streamer(
        key=f"camera-stream-{facing_mode}", # Key must be unique and change to re-render
        mode=WebRtcMode.SENDRECV,
        media_stream_constraints={"video": {"facingMode": facing_mode}, "audio": False},
        video_html_attrs={"autoplay": True, "controls": False, "style": {"width": "100%", "height": "auto"}},
    )
    if webrtc_ctx.video_receiver is None:
        st.spinner("Initializing camera...")

    if st.button("Capture Image"):
        if webrtc_ctx.video_receiver:
            try:
                frame = webrtc_ctx.video_receiver.get_frame(timeout=10) # Increased timeout
                if frame:
                    img = frame.to_image()
                    img_byte_arr = io.BytesIO()
                    # Resize image to reduce file size
                    img.thumbnail((800, 600)) # Resize to max 800x600
                    img.save(img_byte_arr, format='JPEG', quality=85) # Save as JPEG with quality
                    img_byte_arr = img_byte_arr.getvalue()

                    st.session_state.captured_images[photo_name] = img_byte_arr
                    st.success(f"'{photo_name}' captured successfully!")
                    st.rerun()
                else:
                    st.warning("No frame received from the camera. Please try again.")
            except Exception as e:
                st.error(f"Error capturing frame: {e}")
        else:
            st.warning("Camera not ready. Please wait for the video stream to start.")


    # Display captured images
    if st.session_state.captured_images:
        st.subheader("Captured Photos")
        # Create a grid for captured images
        cols = st.columns(3)
        for i, (name, img_data) in enumerate(st.session_state.captured_images.items()):
            with cols[i % 3]:
                st.image(img_data, caption=name, use_column_width=True)

    # Navigation and Download
    st.markdown("<hr>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.button("Back to Form", on_click=set_stage, args=[0])
    with col2:
        if st.session_state.captured_images:
             st.button("Generate and Download Excel", on_click=set_stage, args=[2])


def generate_excel():
    st.header("Download Your Report")
    st.balloons()
    st.success("Your Excel report has been generated!")

    # --- Create Excel Workbook using openpyxl ---
    wb = Workbook()

    # --- Baseband Swap Sheet ---
    ws_baseband = wb.active
    ws_baseband.title = "Baseband Swap"

    # Helper for styling
    def style_cell(cell, text, bold=False, alignment=None):
        cell.value = text
        cell.font = Font(bold=bold)
        if alignment:
            cell.alignment = alignment

    # Form data
    data = st.session_state.baseband_data

    # Title
    ws_baseband.merge_cells('A1:D1')
    style_cell(ws_baseband['A1'], "Gen 4 BB Conversion", bold=True, alignment=Alignment(horizontal='center'))

    # General Info
    style_cell(ws_baseband['A3'], "Antenna Location:", bold=True)
    ws_baseband['B3'].value = data.get('antenna_location')
    if data.get('antenna_location') == 'Other':
        ws_baseband['C3'].value = data.get('antenna_location_other')

    style_cell(ws_baseband['A4'], "Installation:", bold=True)
    ws_baseband['B4'].value = data.get('installation')

    style_cell(ws_baseband['C4'], "Site Name:", bold=True)
    ws_baseband['D4'].value = data.get('site_name')
    style_cell(ws_baseband['C5'], "Contractor:", bold=True)
    ws_baseband['D5'].value = data.get('contractor')
    style_cell(ws_baseband['C6'], "Tech Name:", bold=True)
    ws_baseband['D6'].value = data.get('tech_name')
    style_cell(ws_baseband['C7'], "Date:", bold=True)
    ws_baseband['D7'].value = data.get('date')
    style_cell(ws_baseband['C8'], "Project:", bold=True)
    ws_baseband['D8'].value = data.get('project')

    # Additional Notes
    ws_baseband.merge_cells('A10:D10')
    style_cell(ws_baseband['A10'], "Additional Notes", bold=True, alignment=Alignment(horizontal='center'))
    ws_baseband.merge_cells('A11:D20')
    cell = ws_baseband['A11']
    cell.value = data.get('additional_notes')
    cell.alignment = Alignment(wrap_text=True, vertical='top')


    # --- Photos Sheet ---
    ws_photos = wb.create_sheet(title="Photos")
    row = 1
    col = 1
    for name, img_data in st.session_state.captured_images.items():
        # Add image name
        ws_photos.cell(row=row, column=col, value=name).font = Font(bold=True)

        # Add image
        img = OpenpyxlImage(io.BytesIO(img_data))
        img.width = 300
        img.height = 200
        ws_photos.add_image(img, f"{chr(ord('A') + col -1)}{row + 1}")

        # Adjust row height
        ws_photos.row_dimensions[row + 1].height = 150

        # Move to next position in matrix (2 columns layout)
        col += 1
        if col > 2:
            col = 1
            row += 8 # Adjust spacing for next row of images

    # --- Photo List Sheet ---
    ws_photo_list = wb.create_sheet(title="Photo List")
    for i, photo_name in enumerate(photo_list, 1):
        ws_photo_list.cell(row=i, column=1, value=photo_name)

    # --- Fourth Sheet ---
    wb.create_sheet(title="Sheet4")


    # --- Save workbook to buffer ---
    output = io.BytesIO()
    wb.save(output)
    excel_data = output.getvalue()

    st.download_button(
        label="Download Excel Report",
        data=excel_data,
        file_name=f"{st.session_state.baseband_data.get('site_id', 'report')}_VZW_documentation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.button("Start Over", on_click=set_stage, args=[0])



# --- Main App Logic ---
if st.session_state.stage == 0:
    baseband_swap_form()
elif st.session_state.stage == 1:
    photo_capture_screen()
elif st.session_state.stage == 2:
    generate_excel()

# test
